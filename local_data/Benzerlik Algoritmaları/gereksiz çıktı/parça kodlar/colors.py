from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
import os
import sys


path = os.path.dirname(__file__) + "/cikti2.xlsx"
workbook = load_workbook(filename = path)
sheet = workbook.active
color_scale_rule = ColorScaleRule(start_type="min",
                                   start_color='AA0000',
                                   end_type="max",
                                   end_color='00AA00')

 # Again, let's add this gradient to the star ratings, column "H"
sheet.conditional_formatting.add("E2:E5328", color_scale_rule)
path = os.path.dirname(__file__) + "/cikti_grup_renk.xlsx"
workbook.save(filename = path)